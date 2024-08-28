from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side
from datetime import datetime


class ReportBuilder:
    def __init__(self, ads, statistics):
        self.ads = ads
        self.statistics = statistics

    def build_report(self, filename="report.xlsx"):
        # Generate the current date and append it to the filename
        current_date = datetime.now().strftime("%Y-%m-%d")
        filename_with_date = (
            f"{filename.split('.')[0]}_{current_date}.{filename.split('.')[-1]}"
        )

        # Create a new workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Ads and Statistics"

        # Define some styles for the report
        header_fill = PatternFill(
            start_color="FFC000", end_color="FFC000", fill_type="solid"
        )
        stat_fill = PatternFill(
            start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"
        )
        table_fill = PatternFill(
            start_color="E0FFFF", end_color="E0FFFF", fill_type="solid"
        )
        bold_font = Font(bold=True)

        # Define border style
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Write statistics title
        ws["A1"] = "Ads Statistics"
        ws["A1"].font = bold_font

        # Writing basic statistics (left side of the sheet)
        stats_data = [
            ("Total Ads", self.statistics.total_ads),
            ("Total Views", self.statistics.total_views),
            ("Average Views", self.statistics.average_views),
            ("Oldest Ad Date", self.statistics.oldest_date),
            ("Newest Ad Date", self.statistics.newest_date),
            ("Ads with No Date", self.statistics.ads_with_no_date),
            ("Min Year Built", self.statistics.min_year_built),
            ("Max Year Built", self.statistics.max_year_built),
            ("Average Ads per Day", self.statistics.average_ads_by_day),
        ]

        for row_num, (label, value) in enumerate(stats_data, 2):
            ws[f"A{row_num}"] = label
            ws[f"B{row_num}"] = value

            # Apply styles
            ws[f"A{row_num}"].fill = stat_fill
            ws[f"A{row_num}"].font = bold_font
            ws[f"A{row_num}"].border = thin_border
            ws[f"B{row_num}"].fill = stat_fill
            ws[f"B{row_num}"].border = thin_border

        # Convert dictionaries into sorted tables by value
        currency_dicts = {
            "Average Price per Currency": sorted(
                self.statistics.average_price_per_currency.items(),
                key=lambda x: x[1],
                reverse=True,
            ),
            "Min Price per Currency": sorted(
                self.statistics.min_price_for_each_currency.items(), key=lambda x: x[1]
            ),
            "Max Price per Currency": sorted(
                self.statistics.max_price_for_each_currency.items(),
                key=lambda x: x[1],
                reverse=True,
            ),
            "Currency Usage Count": sorted(
                self.statistics.currency_usage_count.items(),
                key=lambda x: x[1],
                reverse=True,
            ),
        }

        # Write each sorted dictionary into its own table format
        start_row = len(stats_data) + 3
        for title, sorted_currency_dict in currency_dicts.items():
            # Write title
            ws[f"A{start_row}"] = title
            ws[f"A{start_row}"].font = bold_font
            ws[f"A{start_row}"].fill = table_fill

            # Write headers
            ws[f"A{start_row + 1}"] = "Currency"
            ws[f"B{start_row + 1}"] = "Value"
            ws[f"A{start_row + 1}"].font = bold_font
            ws[f"B{start_row + 1}"].font = bold_font
            ws[f"A{start_row + 1}"].fill = header_fill
            ws[f"B{start_row + 1}"].fill = header_fill
            ws[f"A{start_row + 1}"].border = thin_border
            ws[f"B{start_row + 1}"].border = thin_border

            # Write sorted currency data
            for row_num, (currency, value) in enumerate(
                sorted_currency_dict, start_row + 2
            ):
                ws[f"A{row_num}"] = currency
                ws[f"B{row_num}"] = value

                # Apply borders and table styling
                ws[f"A{row_num}"].border = thin_border
                ws[f"B{row_num}"].border = thin_border
                ws[f"A{row_num}"].fill = table_fill
                ws[f"B{row_num}"].fill = table_fill

            # Update start_row for the next table
            start_row = row_num + 2

        # Write num_of_ads_by_day table before the ads list (starting from the 4th column)
        num_of_ads_start_col = 4
        ws[f"{get_column_letter(num_of_ads_start_col)}1"] = "Date"
        ws[f"{get_column_letter(num_of_ads_start_col + 1)}1"] = "Ads Count"
        ws[f"{get_column_letter(num_of_ads_start_col)}1"].font = bold_font
        ws[f"{get_column_letter(num_of_ads_start_col + 1)}1"].font = bold_font
        ws[f"{get_column_letter(num_of_ads_start_col)}1"].fill = header_fill
        ws[f"{get_column_letter(num_of_ads_start_col + 1)}1"].fill = header_fill

        # Write dates and counts in reverse order
        sorted_num_of_ads = sorted(
            self.statistics.num_of_ads_by_day.items(), reverse=True
        )
        for row_num, (date, count) in enumerate(sorted_num_of_ads, 2):
            ws[f"{get_column_letter(num_of_ads_start_col)}{row_num}"] = date
            ws[f"{get_column_letter(num_of_ads_start_col + 1)}{row_num}"] = count

            # Apply borders and styling to the num_of_ads_by_day table
            ws[f"{get_column_letter(num_of_ads_start_col)}{row_num}"].border = (
                thin_border
            )
            ws[f"{get_column_letter(num_of_ads_start_col + 1)}{row_num}"].border = (
                thin_border
            )
            ws[f"{get_column_letter(num_of_ads_start_col)}{row_num}"].fill = table_fill
            ws[f"{get_column_letter(num_of_ads_start_col + 1)}{row_num}"].fill = (
                table_fill
            )

        # Write Ads list starting to the right of num_of_ads_by_day (without 'other' field)
        ads_start_col = num_of_ads_start_col + 3  # Leave a gap between the tables
        ads_headers = [
            "Name",
            "Currency",
            "Price",
            "Favorites Count",
            "Year Built",
            "Ad Date",
            "Views",
        ]

        for col_num, header in enumerate(ads_headers):
            col_letter = get_column_letter(ads_start_col + col_num)
            ws[f"{col_letter}1"] = header

            # Apply styles
            ws[f"{col_letter}1"].fill = header_fill
            ws[f"{col_letter}1"].font = bold_font
            ws[f"{col_letter}1"].border = thin_border

        # Write Ads data next to the num_of_ads_by_day with borders
        for row_num, ad in enumerate(self.ads, 2):
            ws[f"{get_column_letter(ads_start_col)}{row_num}"] = ad.name
            ws[f"{get_column_letter(ads_start_col + 1)}{row_num}"] = ad.currency
            ws[f"{get_column_letter(ads_start_col + 2)}{row_num}"] = ad.price
            ws[f"{get_column_letter(ads_start_col + 3)}{row_num}"] = ad.favorites_count
            ws[f"{get_column_letter(ads_start_col + 4)}{row_num}"] = ad.year_built
            ws[f"{get_column_letter(ads_start_col + 5)}{row_num}"] = ad.ad_date
            ws[f"{get_column_letter(ads_start_col + 6)}{row_num}"] = ad.views

            # Apply borders to ad data
            for col_num in range(ads_start_col, ads_start_col + len(ads_headers)):
                ws[f"{get_column_letter(col_num)}{row_num}"].border = thin_border

        # Save the workbook with the date in the filename
        wb.save(filename_with_date)
        print("Report generated")
